import io
from datetime import datetime
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.platypus import HRFlowable, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlalchemy.orm import Session

from app.models.models import Detection, MonitoringSite, Observation, Species, Survey
from app.services import biodiversity_service, conservation_service, habitat_service


def generate_survey_report_pdf(survey_id: str, db: Session) -> bytes:
    """Generate a detailed PDF report for a single survey expedition."""
    survey = db.query(Survey).filter(Survey.id == survey_id).first()
    if not survey:
        raise ValueError("Survey not found")

    site = db.query(MonitoringSite).filter(MonitoringSite.id == survey.site_id).first()
    observations = db.query(Observation).filter(Observation.survey_id == survey_id).all()

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
    story = []
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle("DocTitle", parent=styles["Title"], fontSize=20, textColor=colors.HexColor("#1B4332"), alignment=0, spaceAfter=6)
    subtitle_style = ParagraphStyle("DocSubTitle", parent=styles["Normal"], fontSize=10, textColor=colors.HexColor("#555555"), spaceAfter=12)
    h2_style = ParagraphStyle("DocH2", parent=styles["Heading2"], fontSize=13, textColor=colors.HexColor("#2D6A4F"), spaceBefore=12, spaceAfter=6)
    body_style = ParagraphStyle("DocBody", parent=styles["Normal"], fontSize=9, leading=12)

    # Header
    story.append(Paragraph("🌿 Wildlife Population Intelligence System", title_style))
    story.append(Paragraph(f"Field Survey Report — Generated on {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}", subtitle_style))
    story.append(HRFlowable(width="100%", thickness=1.5, color=colors.HexColor("#2D6A4F"), spaceAfter=12))

    # Survey Metadata Section
    story.append(Paragraph("Survey Overview", h2_style))
    meta_data = [
        [Paragraph("<b>Survey ID:</b>", body_style), Paragraph(survey.id, body_style), Paragraph("<b>Site Name:</b>", body_style), Paragraph(site.name if site else "N/A", body_style)],
        [Paragraph("<b>Start Date:</b>", body_style), Paragraph(survey.start_date.strftime("%Y-%m-%d"), body_style), Paragraph("<b>Protected Area:</b>", body_style), Paragraph(site.protected_area if site else "N/A", body_style)],
        [Paragraph("<b>Total Observations:</b>", body_style), Paragraph(str(len(observations)), body_style), Paragraph("<b>Device Type:</b>", body_style), Paragraph(site.device_type.value if site else "N/A", body_style)],
        [Paragraph("<b>Notes:</b>", body_style), Paragraph(survey.notes or "None", body_style), "", ""],
    ]
    t_meta = Table(meta_data, colWidths=[110, 160, 110, 160])
    t_meta.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#F8FAF8')),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('PADDING', (0, 0), (-1, -1), 5),
    ]))
    story.append(t_meta)
    story.append(Spacer(1, 12))

    # Observations & Detections List
    story.append(Paragraph("Observations & Species Detections", h2_style))

    det_table_data = [
        [Paragraph("<b>Obs Type</b>", body_style), Paragraph("<b>Captured At</b>", body_style), Paragraph("<b>Species / Label</b>", body_style), Paragraph("<b>Confidence</b>", body_style), Paragraph("<b>Source</b>", body_style)]
    ]

    for obs in observations:
        dets = db.query(Detection).filter(Detection.observation_id == obs.id).all()
        if not dets:
            det_table_data.append([
                Paragraph(obs.observation_type.value.capitalize(), body_style),
                Paragraph(obs.captured_at.strftime("%Y-%m-%d %H:%M"), body_style),
                Paragraph("<i>No detections recorded</i>", body_style),
                "-",
                "-"
            ])
        else:
            for d in dets:
                sp = db.query(Species).filter(Species.id == d.species_id).first() if d.species_id else None
                sp_name = sp.common_name if sp else (d.raw_label or "Unknown")
                conf_pct = f"{int(d.confidence * 100)}%"
                det_table_data.append([
                    Paragraph(obs.observation_type.value.capitalize(), body_style),
                    Paragraph(obs.captured_at.strftime("%Y-%m-%d %H:%M"), body_style),
                    Paragraph(f"<b>{sp_name}</b>" + (f" ({sp.scientific_name})" if sp else ""), body_style),
                    Paragraph(conf_pct, body_style),
                    Paragraph(d.detection_source, body_style)
                ])

    t_dets = Table(det_table_data, colWidths=[80, 110, 200, 70, 80])
    t_dets.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2D6A4F')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')]),
        ('PADDING', (0, 0), (-1, -1), 4),
    ]))
    story.append(t_dets)

    doc.build(story)
    return buffer.getvalue()


def generate_biodiversity_report_pdf(site_id: str, db: Session) -> bytes:
    """Generate a comprehensive Biodiversity & Habitat Score PDF report for a site."""
    site = db.query(MonitoringSite).filter(MonitoringSite.id == site_id).first()
    if not site:
        raise ValueError("Site not found")

    bio_data = biodiversity_service.compute_site_biodiversity(site_id, db)
    hab_data = habitat_service.compute_habitat_score(site_id, db)
    recs = conservation_service.generate_recommendations(site_id, db)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
    story = []
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle("DocTitle", parent=styles["Title"], fontSize=20, textColor=colors.HexColor("#1B4332"), alignment=0, spaceAfter=6)
    subtitle_style = ParagraphStyle("DocSubTitle", parent=styles["Normal"], fontSize=10, textColor=colors.HexColor("#555555"), spaceAfter=12)
    h2_style = ParagraphStyle("DocH2", parent=styles["Heading2"], fontSize=13, textColor=colors.HexColor("#2D6A4F"), spaceBefore=12, spaceAfter=6)
    body_style = ParagraphStyle("DocBody", parent=styles["Normal"], fontSize=9, leading=12)

    # Header
    story.append(Paragraph("🌿 Wildlife Population Intelligence System", title_style))
    story.append(Paragraph(f"Site Biodiversity & Habitat Assessment — {site.name}", subtitle_style))
    story.append(HRFlowable(width="100%", thickness=1.5, color=colors.HexColor("#2D6A4F"), spaceAfter=12))

    # Overview Metrics Table
    story.append(Paragraph("Habitat & Ecological Summary", h2_style))
    summary_table = [
        [Paragraph("<b>Habitat Score:</b>", body_style), Paragraph(f"<b>{hab_data['habitat_score']} / 100 (Grade {hab_data['grade']})</b>", body_style), Paragraph("<b>Classification:</b>", body_style), Paragraph(hab_data['classification'], body_style)],
        [Paragraph("<b>Shannon Index (H'):</b>", body_style), Paragraph(f"{bio_data.get('shannon_index', 0.0):.4f}", body_style), Paragraph("<b>Species Richness:</b>", body_style), Paragraph(f"{bio_data.get('species_richness', 0)} species", body_style)],
        [Paragraph("<b>Protected Area:</b>", body_style), Paragraph(site.protected_area, body_style), Paragraph("<b>Habitat Type:</b>", body_style), Paragraph(site.habitat_type, body_style)],
    ]
    t_sum = Table(summary_table, colWidths=[120, 150, 120, 150])
    t_sum.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#F0FDF4')),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#BBF7D0')),
        ('PADDING', (0, 0), (-1, -1), 5),
    ]))
    story.append(t_sum)
    story.append(Spacer(1, 12))

    # Species Richness Breakdown
    story.append(Paragraph("Observed Species Breakdown", h2_style))
    sp_table_data = [
        [Paragraph("<b>Species Name</b>", body_style), Paragraph("<b>Scientific Name</b>", body_style), Paragraph("<b>Detection Count</b>", body_style), Paragraph("<b>Share (%)</b>", body_style)]
    ]
    tot_dets = max(1, bio_data.get("total_detections", 1))
    for item in bio_data.get("species_breakdown", []):
        pct = round((item["count"] / tot_dets) * 100, 1)
        sp_table_data.append([
            Paragraph(item["species_name"], body_style),
            Paragraph(item.get("scientific_name") or "N/A", body_style),
            Paragraph(str(item["count"]), body_style),
            Paragraph(f"{pct}%", body_style),
        ])

    t_sp = Table(sp_table_data, colWidths=[150, 170, 110, 110])
    t_sp.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E293B')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')]),
        ('PADDING', (0, 0), (-1, -1), 4),
    ]))
    story.append(t_sp)
    story.append(Spacer(1, 12))

    # Conservation Recommendations
    story.append(Paragraph("Priority Conservation Recommendations", h2_style))
    if recs:
        rec_data = [[Paragraph("<b>Priority</b>", body_style), Paragraph("<b>Action Item / Alert</b>", body_style)]]
        for r in recs:
            rec_data.append([
                Paragraph(r["priority"].upper(), body_style),
                Paragraph(r["message"], body_style),
            ])
        t_rec = Table(rec_data, colWidths=[90, 450])
        t_rec.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#991B1B')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#FECACA')),
            ('PADDING', (0, 0), (-1, -1), 5),
        ]))
        story.append(t_rec)
    else:
        story.append(Paragraph("<i>No critical conservation alerts flagged for this site.</i>", body_style))

    doc.build(story)
    return buffer.getvalue()


def export_detections_excel(
    site_id: Optional[str],
    date_from: Optional[str],
    date_to: Optional[str],
    db: Session,
) -> bytes:
    """Generate an Excel workbook of filtered detection records."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Detections Export"

    # Header styling
    header_fill = PatternFill(start_color="1B4332", end_color="1B4332", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    align_center = Alignment(horizontal="center", vertical="center")

    headers = [
        "Detection ID", "Species Common Name", "Scientific Name", "Taxonomic Class",
        "Conservation Status", "Is Endangered", "Confidence (%)", "Count",
        "Detection Source", "Site Name", "Protected Area", "Survey Date", "Timestamp"
    ]
    ws.append(headers)

    for col_num, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center

    query = (
        db.query(Detection, Observation, Survey, MonitoringSite)
        .join(Observation, Detection.observation_id == Observation.id)
        .join(Survey, Observation.survey_id == Survey.id)
        .join(MonitoringSite, Survey.site_id == MonitoringSite.id)
    )

    if site_id and site_id != "all":
        query = query.filter(MonitoringSite.id == site_id)

    if date_from:
        try:
            query = query.filter(Detection.created_at >= datetime.fromisoformat(date_from))
        except ValueError:
            pass

    if date_to:
        try:
            query = query.filter(Detection.created_at <= datetime.fromisoformat(date_to))
        except ValueError:
            pass

    rows = query.order_by(Detection.created_at.desc()).all()

    for det, obs, survey, site in rows:
        species_obj = db.query(Species).filter(Species.id == det.species_id).first() if det.species_id else None

        ws.append([
            det.id,
            species_obj.common_name if species_obj else (det.raw_label or "Unknown"),
            species_obj.scientific_name if species_obj else "N/A",
            species_obj.taxonomic_class if species_obj else "N/A",
            species_obj.conservation_status.value if species_obj else "N/A",
            "Yes" if (species_obj and species_obj.is_endangered) else "No",
            round(det.confidence * 100, 1),
            det.count,
            det.detection_source,
            site.name,
            site.protected_area,
            survey.start_date.strftime("%Y-%m-%d"),
            det.created_at.strftime("%Y-%m-%d %H:%M:%S"),
        ])

    # Auto-adjust column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
